import math as _math
import os as _os
import tempfile as _tmp
import tomllib as _tomllib

import lib_dzne_basetables as _bt
import lib_dzne_math.na as _na
import lib_dzne_seq as _seq
import lib_dzne_tsv as _tsv
import openpyxl as _xl
import pandas as _pd
import tomli_w as _tomli_w


class _File:
    def __init__(self, *, fileDataType, string):
        fileDataType.check_ext(string)
        if not issubclass(fileDataType, FileData):
            raise TypeError()
        if type(string) is not str:
            raise TypeError()
        self._fileDataType = fileDataType
        self._string = string
    @property
    def fileDataType(self):
        return self._fileDataType
    def __str__(self):
        return self._string
    def load(self, /, **kwargs):
        return self._fileDataType.load(file=self._string, **kwargs)
    def save(self, /, data, **kwargs):
        if type(data) is not self._fileDataType:
            raise TypeError()
        return data.save(file=self._string, **kwargs)

class FileData:
    def __init__(self, data):
        if issubclass(type(data), type(self)):
            self.data = data._data
        else:
            self.data = data
    @classmethod
    def file(cls, /, string):
        return _File(fileDataType=cls, string=string)
    @classmethod
    def ext(cls):
        return cls._ext
    @classmethod
    def check_ext(cls, /, file):
        if cls._ext is None:
            return
        if cls._ext == _os.path.splitext(file):
            return
        raise ValueError()
    @classmethod
    def load(cls, /, file, **kwargs):
        if file == "":
            ans = cls._default()
        else:
            cls.check_ext(file)
            ans = cls._load(file=file, **kwargs)
        return cls(ans)
    def save(self, /, file, *, overwrite=False, **kwargs):
        cls = type(self)
        if file == "":
            ans = type(self).drop()
        elif _os.path.exists(file) and not overwrite:
            raise FileExistsError(file)
        else:
            cls.check_ext(file)
            ans = self._save(file=file, **kwargs)
        if ans is not None:
            raise TypeError()
    @classmethod
    def from_file(cls, file, /, *types):
        ext = _os.path.splitext(file)
        for t in types:
            if not issubclass(t, cls):
                raise ValueError()
            if t._ext == ext:
                return t.load(file)
        raise ValueError()
    @classmethod
    def default(cls):
        return cls.load("")
    @property
    def data(self):
        return type(self).clone_data(self._data)
    @data.setter
    def data(self, value):
        self._data = type(self).clone_data(self._data)



class TXTData(FileData):
    _ext = '.txt'
    @classmethod
    def _load(cls, /, file):
        ans = list()
        with open(file, 'r') as s:
            for line in s:
                if not line.endswith('\n'):
                    raise ValueError()
                ans.append(line[:-1])
        return ans
    def _save(self, /, file):
        with open(file, 'w') as s:
            for line in self._data:
                print(line, file=s)
    @staticmethod
    def _default():
        return list()
    @staticmethod
    def clone_data(data):
        x = list()
        for line in data:
            x += str(line).split('\n')
        return x


class TSVData(FileData):
    _ext = '.tsv'
    def __init__(self, dataFrame):
        self.dataFrame = dataFrame
    @classmethod
    def _load(cls, /, file, *, strip=False, **kwargs):
        ans = _tsv.read_DataFrame(file, **kwargs)
        if strip:
            ans = ans.applymap(lambda x: x.strip())
        return cls(ans)
    def _save(self, /, file, *, strip=False):
        ans = self._dataFrame
        if strip:
            ans = ans.applymap(lambda x: x.strip())
        _tsv.write_DataFrame(file, data)
    @staticmethod
    def _default():
        return dict()

    @staticmethod
    def clone_data(data):
        data = _pd.DataFrame(data)
        data = data.copy()
        data = data.applymap(str)
        return data


class BASEData(TSVData):
    @classmethod
    def basetype(cls):
        if cls._ext.startswith("."):
            raise ValueError()
        if cls._ext.startswith("base"):
            raise ValueError()
        return cls._ext[1:-4]
    @classmethod
    def _load(cls, /, file, **kwargs):
        ans = super().load(file, **kwargs).dataFrame
        ans = _bt.table.make(ans, basetype=cls.basetype())
        return ans
    def _save(self, /, file, **kwargs):
        data = _bt.table.make(data, basetype=self.basetype())
        super().save(string, data)
    @classmethod
    def _default(cls):
        return _bt.table.make(basetype=cls.basetype())
    @classmethod
    def from_file(cls, file, /):
        return super().from_file(file, ABASEData, CBASEData, DBASEData, MBASEData, YBASEData)
class ABASEData(BASEData):
    _ext = ".abase"
class CBASEData(BASEData):
    _ext = ".cbase"
class DBASEData(BASEData):
    _ext = ".dbase"
class MBASEData(BASEData):
    _ext = ".mbase"
class YBASEData(BASEData):
    _ext = ".ybase"

class TOMLData(FileData):
    _ext = ".toml"
    def __init__(self, data):
        self.data = data
    def __getitem__(self, key):
        keys = self._getkeys(key)
        ans = self._getitem(*keys)
        ans = self.clone_data(ans)
        return ans
    def __setitem__(self, key, value):
        value = self.clone_data(value)
        keys = self._getkeys(key)
        target = self._getitem(*(keys[:-1]))
        if (type(target) is dict) and (type(keys[-1]) is not str):
            raise TypeError()
        target[keys[-1]] = value
    def __delitem__(self, key):
        keys = self._getkeys(key)
        target = self._getitem(*(keys[:-1]))
        del target[keys[-1]]
    @staticmethod
    def _getkeys(key):
        if type(key) is tuple:
            return key
        else:
            return key,
    def _getitem(self, *keys):
        target = self._data
        for key in keys:
            target = target[key]
        return target
    def items(self, *keys):
        return self[keys].items()
    def append(self, *args):
        *keys, value = args
        self._getitem(*keys).append(value)
    def get(self, *keys, default=None):
        try:
            return self[keys]
        except KeyError:
            return default
    @classmethod
    def clone_data(cls, data):
        if _na.isna(data):
            return float('nan')
        if type(data) in (str, int, bool, float):
            return data
        if data == str(data):
            return str(data)
        try:
            data = dict(data)
        except:
            pass
        else:
            keys = list(data.keys())
            for k in keys:
                if type(k) is not str:
                    raise TypeError()
                data[k] = cls.clone_data(data[k])
            return data
        try:
            return [cls.clone_data(x) for x in data]
        except:
            pass
        raise TypeError()
    @classmethod
    def _load(cls, /, file):
        with open(file, 'rb') as s:
            return _tomllib.load(s)
    def _save(self, /, file):
        with open(file, 'wb') as s:
            _tomli_w.dump(self.data, s)
    @staticmethod
    def _default():
        return dict()

class WorkbookData(FileData):
    _ext = '.xlsx'
    def __init__(self, workbook):
        self.workbook = workbook
    @classmethod
    def _load(cls, /, file):
        return _xl.load_workbook(file)
    def _save(self, /, file):
        self._workbook.save(filename=file)
    @staticmethod
    def _default():
        return _xl.Workbook()
    @classmethod
    def clone_data(workbook):
        with _tmp.TemporaryDirectory() as directory:
            file = _os.path.join(directory, "a" + cls.ext())
            workbook.save(file)
            return _xl.load_workbook(file)
    @staticmethod
    def workbook_from_DataFrames(dataFrames):
        dataFrames = dict(dataFrames)
        if len(dataFrames) == 0:
            return None
        workbook = _xl.Workbook()
        default_sheet = workbook.active
        for table, df in dataFrames.items():
            if default_sheet is None:
                workbook.create_sheet(table)
            else:
                default_sheet.title = table
                default_sheet = None
        for table, df in dataFrames.items():
            columns = list(df.columns)
            for x, column in enumerate(columns):
                workbook[table].cell(row=1, column=x+1).value = column
                for y, v in enumerate(df[column].tolist()):
                    if _pd.isna(v):
                        continue
                    elif (type(v) is float) and (_math.isinf(v)):# is this really needed?
                        value = str(v)
                    else:
                        value = v
                    workbook[table].cell(row=y+2, column=x+1).value = value
        return workbook
    @staticmethod
    def set_cell(*, cell, value):
        """Setting value of cell. """
        if _pd.isna(value):
            value = 'N/A'
        else:
            if type(value) is float:
                if _math.isinf(value):
                    if value < 0:
                        value = '-inf'
                    else:
                        value = '+inf'
            if type(value) not in {str, int, float, bool}:
                raise TypeError(f"The value {value} is of the invalid type {type(value)}! ")
        cell.value = value
        cell.alignment = _xl.styles.Alignment()#horizontal='general')


class SeqReadData(FileData):
    @staticmethod
    def clone_data(data):
        return _seq.SeqRead(read=data)
    @classmethod
    def _load(cls, /, file):
        return _seq.SeqRead(file=file, format=cls._format)
    def _save(self, /, file):
        self._data.save(file=file, format=type(self)._format)
    @classmethod
    def _default(cls):
        return _seq.SeqRead()
    @classmethod
    def from_file(cls, file, /):
        return super().from_file(file, PHDData, ABIData)

class PHDData(SeqRead):
    _ext = '.phd'
    _format = 'phd'
class ABIData(SeqRead):
    _ext = '.ab1'
    _format = 'abi'

